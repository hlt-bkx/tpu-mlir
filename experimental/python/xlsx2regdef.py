import os
import sys
import re
import openpyxl
from fnmatch import fnmatch
from time import gmtime, strftime
from collections import OrderedDict

################### read xlsx ###################

def get_gdma_sheet_name(wb):
    sheet_name = wb.sheetnames
    gdma_name = []
    for name in sheet_name:
        if fnmatch(name, '*DMA_*'):
            gdma_name.append(name)
    return gdma_name

def get_tiu_sheet_name(wb):
    sheet_name = wb.sheetnames
    tiu_name = [n for n in sheet_name]
    for name in sheet_name:
        if name != 'CONV':
            tiu_name.remove(name)
        else:
            break
    return tiu_name

def key_rename(name, keys:list):
    if name not in keys:
        return name
    index = 0
    while(1):
        nm = name + '_' + str(index)
        if nm not in keys:
            break
        index += 1
    return nm

def read_sheet(sheet):
    end = len([v.value for v in sheet["C"] if v.value != None and v.value != "NA"])
    values = sheet["A"][1:] if end == None else sheet["A"][1:end]
    fields = [v.value for v in values if v.value != None]
    fields = [
        i.replace("des_", "").replace("short_", "").split('（')[0].split('(')[0] for i in fields
    ]
    values = sheet["B"][1:end]
    valid = [v.value for v in values if v.value != None]
    assert len(fields) == len(valid)
    reg_order_dict = OrderedDict()
    for k, v in zip(fields, valid):
        reg_order_dict[key_rename(k, reg_order_dict.keys())] = v
    return reg_order_dict

def get_cmd_reg(wb, names, cmd_reg):
    for name in names:
        sheet = wb[name]
        name = name.replace(" ", "")
        name = name.split('（')[0].split('(')[0]
        cmd_reg[name] = read_sheet(sheet)

def read_tiu(wb, cmd_reg):
    tiu_names = get_tiu_sheet_name(wb)
    get_cmd_reg(wb, tiu_names, cmd_reg)

def read_gdma(wb, cmd_reg):
    gdma_names = get_gdma_sheet_name(wb)
    get_cmd_reg(wb, gdma_names, cmd_reg)


################### write reg def cpp ###################

def write_cpp_head(f, chip, file_name):
    cpp_head = f"""// ====- {chip.lower()}RefDef.cpp - {chip.upper()} register definition ==========================
//
// Copyright (C) 2022 Sophgo Technologies Inc.  All rights reserved.
//
// TPU-MLIR is licensed under the 2-Clause BSD License except for the
// third-party components.
//
// ==============================================================================
//
// automatically generated by {__file__}
// time: {strftime('%Y-%m-%d %H:%M:%S', gmtime())}
// this file should not be changed except format.

// reg_def_file: {file_name}

"""
    f.write(cpp_head)
    f.write("#pragma once\n")
    f.write("#include <cstdint>\n")


def write_cmd(f, cmd, reg_def, short=""):
    cmds = ""
    cmd = cmds.join([c[:1].upper()+c[1:] for c in cmd.split("_")])
    bits = sum([s for _, s in reg_def.items()])
    f.write("\nstruct {}{}RegDef {{\n".format(short, cmd))
    f.write("  // {}bits\n".format(bits))
    for reg, size in reg_def.items():
        reg = reg.replace("/", "_")
        f.write("  uint64_t {} : {};\n".format(reg, size))
    f.write("  bool operator==(const {}{}RegDef &rhs) const {{ return memcmp(this, &rhs, sizeof({}{}RegDef)); }}\n".format(short, cmd, short, cmd))
    f.write("};\n")

def write_reg_def(f, file, reg_def):
    reg_def.pop('chip')
    for cmd, reg in reg_def.items():
        if fnmatch(cmd, 's*'):
            cmd = cmd.replace("&s", "_").replace("&", "_")
            write_cmd(f, cmd[1:], reg, "Short")
        else:
            cmd = cmd.replace("&", "_")
            write_cmd(f, cmd, reg)

def gen_reg_def_cpp(reg_def_cpp, cmd_reg_order_dict, *reg_file):
    with open(reg_def_cpp, "w") as f:
        write_cpp_head(f, cmd_reg_order_dict["chip"], reg_file)
        write_reg_def(f, reg_def_cpp, cmd_reg_order_dict)

################### yaml ###################
import yaml
def write_yaml_head(f, tiu_reg, gdma_reg):
    from time import gmtime, strftime
    yaml_head = f"""# ==============================================================================
#
# Copyright (C) 2022 Sophgo Technologies Inc.  All rights reserved.
#
# TPU-MLIR is licensed under the 2-Clause BSD License except for the
# third-party components.
#
# ==============================================================================
#
# automatically generated by {__file__}
# time: {strftime('%Y-%m-%d %H:%M:%S', gmtime())}
# this file should not be changed except format.

# tiu_reg_fn: {tiu_reg}
# dma_reg_fn: {gdma_reg}

"""
    f.write(yaml_head)

def ordered_yaml_dump(data, stream=None, Dumper=yaml.SafeDumper, **kwds):
    class OrderedDumper(Dumper):
        pass

    def _dict_representer(dumper, data):
        return dumper.represent_mapping(
            yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
            data.items())

    OrderedDumper.add_representer(OrderedDict, _dict_representer)
    return yaml.dump(data, stream, OrderedDumper, **kwds)

def write_to_yaml(yaml_file, cmd_reg_order_dict, tiu_reg, gdma_reg):
    with open(yaml_file, "w") as f:
        write_yaml_head(f, tiu_reg, gdma_reg)
        ordered_yaml_dump(cmd_reg_order_dict, f)

def ordered_yaml_load(yaml_path, Loader=yaml.Loader,
                      object_pairs_hook=OrderedDict):
    class OrderedLoader(Loader):
        pass

    def construct_mapping(loader, node):
        loader.flatten_mapping(node)
        return object_pairs_hook(loader.construct_pairs(node))

    OrderedLoader.add_constructor(
        yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
        construct_mapping)
    with open(yaml_path) as stream:
        return yaml.load(stream, OrderedLoader)

def yaml_to_cpp(reg_def_cpp, reg_def_yaml):
    reg_yaml = ordered_yaml_load(reg_def_yaml)
    gen_reg_def_cpp(reg_def_cpp, reg_yaml, reg_def_yaml)

################### main ###################

if __name__ == "__main__":
    if (len(sys.argv) < 5):
        print('usage: python3 ' + sys.argv[0] + ' chip RegDef.h tiu_reg.xlsx gdma_reg.xlsx')
        exit(-1)
    TIU_REG=sys.argv[3]
    GDMA_REG=sys.argv[4]
    REG_DEF_CPP = sys.argv[2]
    CHIP = sys.argv[1].upper()
    cmd_reg_order_dict = OrderedDict()
    cmd_reg_order_dict["chip"] = CHIP
    tiu_wb = openpyxl.load_workbook(TIU_REG, data_only=True)
    read_tiu(tiu_wb, cmd_reg_order_dict)
    gdma_wb = openpyxl.load_workbook(GDMA_REG, data_only=True)
    read_gdma(gdma_wb, cmd_reg_order_dict)
    gen_reg_def_cpp(REG_DEF_CPP, cmd_reg_order_dict, TIU_REG, GDMA_REG)
    # write_to_yaml("regdef.yaml", cmd_reg_order_dict, TIU_REG, GDMA_REG)
    # yaml_to_cpp(REG_DEF_CPP, "regdef.yaml")
